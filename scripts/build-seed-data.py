#!/usr/bin/env python3
"""Build real seed-data JSON from the KP «Послуга» spreadsheets.

Source (committed under prisma/seed-data/source/):
  grafiky-vyvezennya-kp-posluha.xlsx  — real TPV/BULK collection schedules,
                                        container sites, vehicles.

Outputs (committed under prisma/seed-data/, consumed by prisma/seed.ts):
  vehicles.json            — 9 real trucks (name + plate)
  collection-points.json   — 61 real container sites with fractions + coords
  streets.json             — 248 distinct streets, each with an aggregated
                             weekday/time schedule and (where matched) a
                             primary container site
  package-schedules.json   — curbside schedule groups for streets without a
                             matched container site

Run:  pip install openpyxl && python3 scripts/build-seed-data.py
Re-run whenever the source spreadsheet changes.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "prisma" / "seed-data" / "source" / "grafiky-vyvezennya-kp-posluha.xlsx"
OUT_DIR = ROOT / "prisma" / "seed-data"

OPERATOR_NAME = "КП «Послуга»"

# Pryluky bounding box (WGS84). Rows geocoded outside this are dropped as
# obviously-wrong geocoder hits (other villages in Chernihiv Oblast).
BBOX = (50.53, 50.66, 32.28, 32.50)

WEEKDAY_TO_NUM = {
    "понеділок": 1,
    "вівторок": 2,
    "середа": 3,
    "четвер": 4,
    "пятниця": 5,
    "субота": 6,
    "неділя": 7,
}


def norm_weekday(s: str) -> int | None:
    key = re.sub(r"[^а-яіїєґ]", "", str(s).strip().lower())
    return WEEKDAY_TO_NUM.get(key)


def in_bbox(lat, lng) -> bool:
    try:
        lat = float(lat)
        lng = float(lng)
    except (TypeError, ValueError):
        return False
    return BBOX[0] <= lat <= BBOX[1] and BBOX[2] <= lng <= BBOX[3]


def clean_time(v) -> str | None:
    """'9:00' / '13:05' / time-object -> 'HH:MM' 24h."""
    if v is None or v == "":
        return None
    if hasattr(v, "hour"):
        return f"{v.hour:02d}:{v.minute:02d}"
    m = re.match(r"^\s*(\d{1,2}):(\d{2})", str(v))
    if not m:
        return None
    return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"


def street_of(address: str) -> str:
    """Street part of a site address: 'вул. Переяславська, 3' -> 'вул. Переяславська'."""
    return re.split(r",", str(address), 1)[0].strip()


def norm_street_key(name: str) -> str:
    """Loose key for matching street-name variants. Keeps a type token so a
    street and a lane of the same name ("вул. Шевченка" vs "пров. Шевченка")
    stay distinct, while casing / prefix-omission variants collapse."""
    s = str(name).lower().strip()
    if re.search(r"пров(ул)?\.?|провулок", s):
        kind = "p"
    elif re.search(r"в['’]?їзд|в-д", s):
        kind = "e"
    elif re.search(r"площ|пл\.", s):
        kind = "s"
    else:
        kind = "v"  # вулиця / bare name
    s = re.sub(r"вул(иця)?\.?|пров(ул)?\.?|провулок|в['’]?їзд|в-д|площа|пл\.", "", s)
    s = re.sub(r"[^а-яіїєґ0-9]", "", s)
    return f"{kind}:{s}"


def parse_int_cell(v) -> int:
    """'3' / '3*' / 3 -> 3 ; None -> 0."""
    if v is None:
        return 0
    m = re.match(r"\s*(\d+)", str(v))
    return int(m.group(1)) if m else 0


def periodicity_for(day_count: int) -> str:
    if day_count >= 5:
        return "DAILY"
    if day_count == 2:
        return "TWICE_WEEKLY"
    if day_count == 1:
        return "WEEKLY"
    return "TWICE_WEEKLY"  # 3-4 days/week: enum has no 3x — array carries the truth


# --------------------------------------------------------------------------- #

wb = openpyxl.load_workbook(SRC, data_only=True)


# ---- vehicles -------------------------------------------------------------- #

TITLE_RE = re.compile(r"автомобіль\s+(.+?),\s*реєстр\.?\s*№?\s*(.+?)\s*$")

vehicles: list[dict] = []
seen_plates: set[str] = set()
for sn in wb.sheetnames:
    if not re.match(r"^М\d", sn):
        continue
    title = next(wb[sn].iter_rows(values_only=True))[0] or ""
    m = TITLE_RE.search(title)
    if not m:
        continue
    name = m.group(1).strip()
    plate = m.group(2).strip().rstrip("*").strip()
    if plate in seen_plates:
        continue
    seen_plates.add(plate)
    is_bulk = "великогабаритн" in title.lower()
    vehicles.append(
        {
            "plateNumber": plate,
            "vehicleType": f"{name} (сміттєвоз)" if not is_bulk else f"{name} (для ВГВ)",
            "vehicleName": name,
            # Real body volume not present in the source; 0 = не вказано.
            "capacityM3": 12 if is_bulk else 0,
            "fuelNormLPer100km": None,
        }
    )

# Map a free-text vehicle string from the schedule to a known plate.
name_to_plate: dict[str, str] = {}
for v in vehicles:
    name_to_plate[norm_street_key(v["vehicleName"])] = v["plateNumber"]


def match_vehicle(cell: str | None) -> str | None:
    if not cell:
        return None
    first = str(cell).split(",")[0].strip()
    return name_to_plate.get(norm_street_key(first))


# ---- ВГВ (bulk) addresses ----------------------------------------------- #

def addr_key(street, house) -> str:
    hk = re.sub(r"[^0-9]", "", str(house or ""))
    return norm_street_key(street) + "|" + hk


bulk_ws = wb["М2_ВГВ_РеноЛендер_0589"]
bulk_rows = list(bulk_ws.iter_rows(values_only=True))
bulk_keys: set[str] = set()  # "<streetkey>|<housedigits>" with ВГВ service
for r in bulk_rows[4:]:
    if not r[1] or str(r[1]).startswith("ВСЬОГО"):
        continue
    bulk_keys.add(addr_key(r[1], r[2]))


# ---- street schedules (aggregated per street from Зведена_по_адресах) --- #

zv = wb["Зведена_по_адресах"]
zrows = list(zv.iter_rows(values_only=True))

# Merge rows by a loose street key so casing / punctuation variants
# ("В/м 12" vs "в/м 12") collapse into one Street.
agg: dict[str, dict] = {}
name_votes: dict[str, dict] = defaultdict(lambda: defaultdict(int))
for r in zrows[4:]:
    name = r[0]
    if not name:
        continue
    name = str(name).strip()
    key = norm_street_key(name)
    if not key:
        continue
    name_votes[key][name] += 1
    a = agg.setdefault(
        key,
        {
            "name": name,
            "daytimes": set(),  # (weekday_num, "HH:MM")
            "vehicleCells": [],
            "linkedSiteNames": set(),
            "coord": None,
        },
    )
    for col in range(2, 9):  # Пн..Нд
        t = r[col]
        if t in (None, ""):
            continue
        wd = col - 1  # col 2 -> Monday(1)
        for piece in str(t).split(","):
            hhmm = clean_time(piece)
            if hhmm:
                a["daytimes"].add((wd, hhmm))
    if r[11]:
        a["vehicleCells"].append(str(r[11]))
    site_link = r[19]
    if site_link:
        a["linkedSiteNames"].add(str(site_link).strip())
    if a["coord"] is None and in_bbox(r[17], r[18]):
        a["coord"] = [float(r[17]), float(r[18])]


def _schedule_from_agg(a: dict) -> dict | None:
    if not a or not a["daytimes"]:
        return None
    daytimes = sorted(a["daytimes"])
    days = sorted({wd for wd, _ in daytimes})
    time_from = min(t for _, t in daytimes)
    vehicle_plate = None
    for cell in a["vehicleCells"]:
        vehicle_plate = match_vehicle(cell)
        if vehicle_plate:
            break
    return {
        "daysOfWeek": days,
        "timeFrom": time_from,
        "periodicity": periodicity_for(len(days)),
        "vehiclePlate": vehicle_plate,
    }


def street_schedule_for(address: str) -> dict | None:
    """Best street-level schedule for a container site, matched by street name
    (handles 'Перехрестя A - B' by trying each part)."""
    first = re.split(r",", str(address), 1)[0]
    first = re.sub(r"\(.*?\)", " ", first)  # drop "(АТБ)", "(3-й магазин)" тощо
    for part in re.split(r"\s[-–—]\s|перехрестя", first, flags=re.IGNORECASE):
        part = part.strip()
        if not part:
            continue
        sched = _schedule_from_agg(agg.get(norm_street_key(part)))
        if sched:
            return sched
    return None


# ---- collection points (61 real sites) --------------------------------- #

sites_ws = wb["Координати майданчиків"]
srows = list(sites_ws.iter_rows(values_only=True))
points: list[dict] = []
site_by_street: dict[str, dict] = {}
site_by_address: dict[str, dict] = {}

FRACTIONS = [
    (3, "PAPER", 240),
    (4, "GLASS", 240),
    (5, "PLASTIC", 240),
    (6, "MIXED", 1100),
]

for r in srows[3:]:
    num, address, obj_type = r[0], r[1], r[2]
    if not address or str(address).strip().startswith("Всього"):
        continue
    if str(address).strip().rstrip(":").lower() in ("всього", "разом"):
        continue
    lat, lng = r[7], r[8]
    try:
        lat, lng = float(lat), float(lng)
    except (TypeError, ValueError):
        print(f"  ! site skipped (no coordinates): {address}")
        continue
    if not in_bbox(lat, lng):
        print(f"  ! site out of bbox, kept with raw coords: {address} ({lat},{lng})")
    containers = []
    for col, cat, vol in FRACTIONS:
        qty = parse_int_cell(r[col])
        if qty > 0:
            containers.append(
                {"wasteCategory": cat, "volumeLiters": vol, "quantity": qty}
            )
    if not containers:
        containers.append(
            {"wasteCategory": "MIXED", "volumeLiters": 1100, "quantity": 1}
        )
    parts = re.split(r",", str(address), 1)
    st_key = norm_street_key(parts[0])
    house = parts[1] if len(parts) > 1 else ""
    is_bulk = addr_key(parts[0], house) in bulk_keys
    if is_bulk:
        # Майданчики з послугою ВГВ мають бункер великогабаритних відходів
        # (Методика роздільного збирання: контейнери ВГВ місткістю ≥ 5 м³).
        containers.append(
            {"wasteCategory": "BULK", "volumeLiters": 5000, "quantity": 1}
        )
    point = {
        "address": str(address).strip(),
        "lat": lat,
        "lng": lng,
        "operatorName": OPERATOR_NAME,
        "isBulkWasteSite": is_bulk,
        "internalNotes": (
            None
            if obj_type == "Контейнерний майданчик"
            else (str(obj_type).strip() if obj_type else None)
        ),
        "containers": containers,
        # Графік вивезення для КОЖНОГО майданчика — з розкладу вулиці, на якій
        # він стоїть (не лише для тих, що співпали з CONTAINER-вулицею).
        "schedule": street_schedule_for(address),
    }
    points.append(point)
    site_by_address[str(address).strip()] = point
    # first site wins as the street's default primary
    site_by_street.setdefault(st_key, point)


# ---- streets --------------------------------------------------------------- #

streets_out: list[dict] = []
package_groups: dict[tuple, list[str]] = defaultdict(list)
matched_container = 0
matched_package = 0
no_schedule = 0

for key, a in sorted(agg.items(), key=lambda kv: kv[1]["name"].lower()):
    # canonical display name = most frequent raw variant, ties -> longest
    name = max(name_votes[key].items(), key=lambda kv: (kv[1], len(kv[0])))[0]
    daytimes = sorted(a["daytimes"])
    days = sorted({wd for wd, _ in daytimes})
    time_from = min((t for _, t in daytimes), default=None)
    periodicity = periodicity_for(len(days))
    vehicle_plate = None
    for cell in a["vehicleCells"]:
        vehicle_plate = match_vehicle(cell)
        if vehicle_plate:
            break

    # match a real container site
    site = None
    for sn in a["linkedSiteNames"]:
        site = site_by_address.get(sn)
        if site:
            break
    if site is None:
        site = site_by_street.get(norm_street_key(name))

    entry = {
        "name": name,
        "collectionMethod": "CONTAINER" if site else "PACKAGE",
        "primaryPointAddress": site["address"] if site else None,
    }

    if not days or not time_from:
        no_schedule += 1
        entry["schedule"] = None
        entry["collectionMethod"] = "CONTAINER" if site else "PACKAGE"
        streets_out.append(entry)
        continue

    if site:
        matched_container += 1
        entry["schedule"] = {
            "daysOfWeek": days,
            "timeFrom": time_from,
            "periodicity": periodicity,
            "vehiclePlate": vehicle_plate,
        }
    else:
        matched_package += 1
        entry["schedule"] = None
        package_groups[(tuple(days), time_from, periodicity)].append(name)

    streets_out.append(entry)

packages_out = [
    {
        "daysOfWeek": list(days),
        "timeFrom": time_from,
        "periodicity": periodicity,
        "streetNames": sorted(names),
    }
    for (days, time_from, periodicity), names in sorted(
        package_groups.items(), key=lambda kv: (kv[0][1], kv[0][0])
    )
]


# ---- write --------------------------------------------------------------- #

for v in vehicles:
    v.pop("vehicleName", None)

(OUT_DIR / "vehicles.json").write_text(
    json.dumps(vehicles, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
)
(OUT_DIR / "collection-points.json").write_text(
    json.dumps(points, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
)
(OUT_DIR / "streets.json").write_text(
    json.dumps(streets_out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
)
(OUT_DIR / "package-schedules.json").write_text(
    json.dumps(packages_out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
)

print("\n=== build-seed-data summary ===")
print(f"vehicles                : {len(vehicles)}")
print(f"collection points       : {len(points)}")
print(f"  bulk-waste sites       : {sum(1 for p in points if p['isBulkWasteSite'])}")
print(f"  total containers        : {sum(len(p['containers']) for p in points)}")
print(f"  with a schedule        : {sum(1 for p in points if p['schedule'])}")
print(f"  WITHOUT a schedule     : {[p['address'] for p in points if not p['schedule']]}")
print(f"streets                 : {len(streets_out)}")
print(f"  CONTAINER (site match) : {sum(1 for s in streets_out if s['collectionMethod']=='CONTAINER')}")
print(f"  PACKAGE (curbside)     : {sum(1 for s in streets_out if s['collectionMethod']=='PACKAGE')}")
print(f"  with a schedule        : {sum(1 for s in streets_out if s['schedule'] or s['collectionMethod']=='PACKAGE')}")
print(f"  no schedule at all     : {no_schedule}")
print(f"package schedule groups : {len(packages_out)}")
missing_veh = sum(
    1 for s in streets_out if s["schedule"] and not s["schedule"]["vehiclePlate"]
)
print(f"CONTAINER schedules missing a vehicle match: {missing_veh}")
